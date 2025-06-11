import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('TkAgg')  # Set interactive backend
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import openpyxl
import os

# Ensure output directory exists
os.makedirs('output', exist_ok=True)

# Automatically generate Bitcoin halving dates
def generate_halving_dates(start_date, end_date):
    halving_dates = []
    first_halving = pd.to_datetime('2012-11-28')  # First halving
    halving_interval = timedelta(days=1460)  # Approximately 4 years (210,000 blocks)
    current_halving = first_halving
    
    # Generate halving dates until exceeding end_date
    while current_halving <= pd.to_datetime(end_date) + timedelta(days=1460):
        halving_dates.append(current_halving)
        current_halving += halving_interval
    
    print(f"Generated halving dates: {[d.strftime('%Y-%m-%d') for d in halving_dates]}")
    return halving_dates

# Fetch Bitcoin historical price data
def get_bitcoin_data(start_date, end_date):
    try:
        btc = yf.download('BTC-USD', start=start_date, end=end_date, auto_adjust=False)
        if btc.empty:
            raise ValueError("No data retrieved from yfinance.")
        print(f"Data retrieved: {len(btc)} rows from {start_date} to {end_date}")
        return btc
    except Exception as e:
        print(f"Error downloading data: {e}")
        return pd.DataFrame()

# Calculate Exponential Moving Average (EMA)
def calculate_ema(data, period):
    return data['Close'].ewm(span=period, adjust=False).mean()

# Calculate Relative Strength Index (RSI)
def calculate_rsi(data, period=14):
    delta = data['Close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

# Detect death cross
def detect_death_cross(ema_short, ema_long):
    death_cross = (ema_short.shift(1) > ema_long.shift(1)) & (ema_short < ema_long)
    return death_cross

# Check if within approximately 500 days before a halving
def is_near_halving_low(current_date, halving_dates, days_window=30):
    target_date = pd.to_datetime(current_date)
    for halving in halving_dates:
        low_date = halving - timedelta(days=500)
        if abs((target_date - low_date).days) <= days_window:
            return True
    return False

# Check if in Q2 (April-June)
def is_in_q2(date):
    return date.month in [4, 5, 6]

# Check if in the bottom region of the rainbow chart (price below 80% of 200-day EMA)
def is_in_rainbow_bottom(close_price, ema_200, threshold=0.8):
    if pd.isna(close_price) or pd.isna(ema_200):
        return False
    return close_price < ema_200 * threshold

# Main function: Generate buy signals
def find_buy_signals(start_date, end_date, halving_dates, min_conditions=2):
    # Fetch data
    data = get_bitcoin_data(start_date, end_date)
    if data.empty:
        print("No data available. Exiting.")
        return pd.DataFrame()
    
    # Calculate EMA and RSI
    data['EMA_100'] = calculate_ema(data, 100)
    data['EMA_200'] = calculate_ema(data, 200)
    data['RSI'] = calculate_rsi(data, 14)
    
    # Detect death cross
    data['Death_Cross'] = detect_death_cross(data['EMA_100'], data['EMA_200'])
    
    # Initialize buy signal column
    data['Buy_Signal'] = False
    
    # Iterate through data to generate buy signals
    for i in range(14, len(data)):  # Start from 14 to ensure RSI has enough data
        current_date = data.index[i]
        close_price = float(data['Close'].iloc[i])
        ema_200 = float(data['EMA_200'].iloc[i])
        rsi = float(data['RSI'].iloc[i] if not pd.isna(data['RSI'].iloc[i]) else 0)
        
        # Buy conditions
        near_halving = is_near_halving_low(current_date, halving_dates)
        recent_death_cross = bool(data['Death_Cross'].iloc[i-10:i].any() if i >= 10 else False)
        in_rainbow_bottom = is_in_rainbow_bottom(close_price, ema_200, threshold=0.8)
        in_q2 = is_in_q2(current_date)
        rsi_oversold = bool(rsi < 30 if not pd.isna(rsi) else False)
        
        # Store conditions for today (only for the last row)
        if i == len(data) - 1:
            today_conditions = {
                'near_halving': near_halving,
                'recent_death_cross': recent_death_cross,
                'in_rainbow_bottom': in_rainbow_bottom,
                'in_q2': in_q2,
                'rsi_oversold': rsi_oversold
            }
        
        # Calculate the number of conditions met
        conditions_met = sum([near_halving, recent_death_cross, in_rainbow_bottom, in_q2, rsi_oversold])
        
        # Buy signal: At least min_conditions are met
        if conditions_met >= min_conditions:
            data.iloc[i, data.columns.get_loc('Buy_Signal')] = True
    
    # Check if today is a buy point
    today = data.index[-1]
    is_today_buy = data['Buy_Signal'].iloc[-1]
    print(f"\nIs today ({today.date()}) a buy point? {'Yes' if is_today_buy else 'No'}")
    if is_today_buy:
        print(f"Conditions met for today ({sum(today_conditions.values())}/{min_conditions}):")
        for condition, value in today_conditions.items():
            print(f"  {condition}: {value}")
    else:
        print(f"Conditions met for today ({sum(today_conditions.values())}/{min_conditions}):")
        for condition, value in today_conditions.items():
            print(f"  {condition}: {value}")
    
    return data

# Visualize results
def plot_data(data, save_plot=False, plot_file='output/bitcoin_plot.png'):
    if data.empty:
        print("No data to plot.")
        return
    
    print("Generating plot...")
    plt.figure(figsize=(14, 10))
    
    # Price and EMA plot
    plt.subplot(2, 1, 1)
    plt.plot(data.index, data['Close'], label='BTC Price', color='blue')
    plt.plot(data.index, data['EMA_100'], label='100-day EMA', color='orange')
    plt.plot(data.index, data['EMA_200'], label='200-day EMA', color='green')
    
    # Mark buy signals
    buy_signals = data[data['Buy_Signal']]
    plt.scatter(buy_signals.index, buy_signals['Close'], color='red', marker='^', s=100, label='Buy Signal')
    
    plt.title('Bitcoin Price with Buy Signals')
    plt.xlabel('Date')
    plt.ylabel('Price (USD)')
    plt.legend()
    plt.grid()
    
    # RSI plot
    plt.subplot(2, 1, 2)
    plt.plot(data.index, data['RSI'], label='RSI', color='purple')
    plt.axhline(y=30, color='red', linestyle='--', label='Oversold (30)')
    plt.title('Relative Strength Index (RSI)')
    plt.xlabel('Date')
    plt.ylabel('RSI')
    plt.legend()
    plt.grid()
    
    plt.tight_layout()
    
    if save_plot:
        plt.savefig(plot_file)
        print(f"Plot saved to {plot_file}")
        plt.close()
    else:
        plt.show()
        print("Plot displayed.")

# Main execution flow
if __name__ == "__main__":
    # Custom parameters
    start_date = '2018-01-01'
    end_date = datetime.today().strftime('%Y-%m-%d')  # Use today's date (2025-06-11)
    min_conditions = 2  # Require at least 2 conditions to be met
    excel_file = 'output/bitcoin_buy_signals.xlsx'  # Excel output filename
    
    # Automatically generate halving dates
    halving_dates = generate_halving_dates(start_date, end_date)
    
    # Perform analysis
    data = find_buy_signals(start_date, end_date, halving_dates, min_conditions)
    
    # Process buy signals and calculate return rate
    buy_signals = data[data['Buy_Signal']]
    if not buy_signals.empty:
        print("\nPotential Buy Signals:")
        print(buy_signals[['Close', 'EMA_100', 'EMA_200', 'RSI']])
        
        # Calculate average buy price and return rate
        today_close = data['Close'].iloc[-1]  # Today's closing price
        if isinstance(today_close, pd.Series):
            today_close = today_close.item()  # Ensure scalar
        avg_buy_price = buy_signals['Close'].mean()  # Average buy price
        if isinstance(avg_buy_price, pd.Series):
            avg_buy_price = avg_buy_price.item()  # Ensure scalar
        return_rate = ((today_close - avg_buy_price) / avg_buy_price) * 100 if avg_buy_price > 0 else 0
        total_buy_points = len(buy_signals)
        
        print(f"\nInvestment Summary (as of {end_date}):")
        print(f"Total Buy Points: {total_buy_points}")
        print(f"Average Buy Price: ${avg_buy_price:,.2f}")
        print(f"Today's Price: ${today_close:,.2f}")
        print(f"Total Return Rate: {return_rate:.2f}%")
        
        # Save to Excel
        try:
            # Save buy signals data
            buy_signals[['Close', 'EMA_100', 'EMA_200', 'RSI']].to_excel(excel_file, sheet_name='Buy_Signals', index=True)
            
            # Use openpyxl to add summary data
            wb = openpyxl.load_workbook(excel_file)
            ws = wb['Buy_Signals']
            summary_row = len(buy_signals) + 3  # Leave one blank row below data
            ws.cell(row=summary_row, column=1).value = "Investment Summary"
            ws.cell(row=summary_row + 1, column=1).value = "Total Buy Points"
            ws.cell(row=summary_row + 1, column=2).value = total_buy_points
            ws.cell(row=summary_row + 2, column=1).value = "Average Buy Price"
            ws.cell(row=summary_row + 2, column=2).value = avg_buy_price
            ws.cell(row=summary_row + 3, column=1).value = "Today's Price"
            ws.cell(row=summary_row + 3, column=2).value = today_close
            ws.cell(row=summary_row + 4, column=1).value = "Total Return Rate (%)"
            ws.cell(row=summary_row + 4, column=2).value = return_rate / 100  # Excel displays percentage as decimal
            wb.save(excel_file)
            print(f"Buy signals and investment summary saved to {excel_file}")
        except Exception as e:
            print(f"Error saving to Excel: {e}")
    else:
        print("\nNo buy signals found.")
    
    # Visualize (save to file to avoid hanging)
    plot_data(data, save_plot=True, plot_file='output/bitcoin_plot.png')
    print("Script completed.")